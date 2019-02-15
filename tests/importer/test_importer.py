import os
import unittest
from unittest import TestCase

from mock import MagicMock, patch
from openpyxl import Workbook

from ingest.importer.conversion.metadata_entity import MetadataEntity
from ingest.importer.importer import WorksheetImporter, WorkbookImporter, \
    IdentifiableWorksheetImporter
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook

BASE_PATH = os.path.dirname(__file__)

HEADER_ROW = 4


def _create_single_row_worksheet(worksheet_data: dict):
    workbook = Workbook()
    worksheet = workbook.create_sheet()

    for column, data in worksheet_data.items():
        key, value = data
        worksheet[f'{column}{HEADER_ROW}'] = key
        worksheet[f'{column}6'] = value

    return worksheet


class WorkbookImporterTest(TestCase):

    @patch('ingest.importer.importer.IdentifiableWorksheetImporter')
    def test_do_import(self, worksheet_importer_constructor):
        # given:
        template_mgr = MagicMock(name='template_manager')
        worksheet_importer = WorksheetImporter(template_mgr)
        worksheet_importer_constructor.return_value = worksheet_importer

        # and:
        user = MetadataEntity(concrete_type='user', domain_type='user', object_id=1,
                              content={'user_name': 'jdelacruz'})
        worksheet_importer.do_import = MagicMock(side_effect=[[user], []])

        # and:
        workbook = Workbook()
        workbook.create_sheet('Users')
        workbook.create_sheet('Items')

        # and: remove the magical default worksheet
        default_sheet = workbook.get_sheet_by_name('Sheet')
        workbook.remove(default_sheet)

        # when:
        ingest_workbook = IngestWorkbook(workbook)
        workbook_importer = WorkbookImporter(template_mgr)
        workbook_json = workbook_importer.do_import(ingest_workbook)

        # then:
        self.assertIsNotNone(workbook_json)

        # and:
        user_map = workbook_json.get('user')
        self.assertIsNotNone(user_map)
        self.assertIn(user.object_id, user_map.keys())
        self.assertEqual({'user_name': 'jdelacruz'}, user_map.get(1)['content'])

    @patch('ingest.importer.importer.IdentifiableWorksheetImporter')
    @unittest.skip
    def test_old_do_import(self, worksheet_importer_constructor):
        # given: set up template manager
        key_label_map = {
            'Project': 'project',
            'Cell Suspension': 'cell_suspension'
        }

        domain_entity_map = {
            'project': 'project',
            'cell_suspension': 'biomaterial'
        }

        mock_template_manager = MagicMock()
        mock_template_manager.get_concrete_type = lambda key: key_label_map.get(key)
        mock_template_manager.get_domain_type = lambda key: domain_entity_map.get(key)

        # and: set up worksheet importer
        worksheet_importer = IdentifiableWorksheetImporter(mock_template_manager)
        expected_json = self._fake_worksheet_import(worksheet_importer, mock_template_manager)

        # and: set up workbook
        workbook = Workbook()
        ingest_workbook = IngestWorkbook(workbook)
        schema_list = self._mock_get_schemas(ingest_workbook)
        self._mock_importable_worksheets(ingest_workbook, workbook)

        # and: mock WorksheetImporter constructor
        worksheet_importer_constructor.return_value = worksheet_importer
        workbook_importer = WorkbookImporter(mock_template_manager)
        workbook_importer.import_or_reference_project = MagicMock()

        # when:
        workbook_output = workbook_importer.do_import(ingest_workbook)

        # then:
        self.assertEqual(2, len(list(workbook_output.keys())))
        self.assertEqual(['project', 'biomaterial'], list(workbook_output.keys()))
        self.assertEqual(2, len(list(workbook_output['project'].keys())))
        self.assertEqual(expected_json['project'], workbook_output['project'])
        self.assertEqual(expected_json['biomaterial'], workbook_output['biomaterial'])

    def _mock_get_schemas(self, ingest_workbook):
        schema_base_url = 'https://schema.humancellatlas.org'
        schema_list = [
            f'{schema_base_url}/type/project',
            f'{schema_base_url}/type/biomaterial'
        ]
        ingest_workbook.get_schemas = MagicMock(return_value=schema_list)
        return schema_list

    def _mock_importable_worksheets(self, ingest_workbook, workbook):
        project_worksheet = workbook.create_sheet('Project')
        cell_suspension_worksheet = workbook.create_sheet('Cell Suspension')
        ingest_workbook.importable_worksheets = MagicMock(return_value=[
            project_worksheet, cell_suspension_worksheet
        ])

    def _fake_worksheet_import(self, worksheet_importer: WorksheetImporter, mock_template_manager):
        projects = {
            'project 1': {'short_name': 'project 1', 'description': 'first project'},
            'project 2': {'short_name': 'project 2', 'description': 'second project'}
        }

        cell_suspensions = {
            'cell_suspension_101': {'biomaterial_id': 'cell_suspension_101', 'biomaterial_name': 'cell suspension'}
        }

        worksheet_iterator = iter([projects, cell_suspensions])
        worksheet_importer.do_import = (
            lambda __, tm: worksheet_iterator.__next__() if tm is mock_template_manager else []
        )

        return {
            'project': projects,
            'biomaterial': cell_suspensions
        }


class WorksheetImporterTest(TestCase):

    def test_do_import(self):
        # given:
        row_template = MagicMock('row_template')

        # and:
        john_doe = MetadataEntity(object_id='profile_1')
        emma_jackson = MetadataEntity(object_id='profile_2')
        row_template.do_import = MagicMock('import_row', side_effect=[john_doe, emma_jackson])

        # and:
        mock_template_manager = MagicMock('template_manager')
        mock_template_manager.create_row_template = MagicMock(return_value=row_template)
        mock_template_manager.get_header_row = MagicMock(return_value=['header1', 'header2'])
        mock_template_manager.get_concrete_type = MagicMock(return_value='concrete_entity')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('user_profile')
        worksheet['A6'] = 'john'
        worksheet['A7'] = 'emma'

        # when:
        worksheet_importer = WorksheetImporter(mock_template_manager)
        profiles = worksheet_importer.do_import(worksheet)

        # then:
        self.assertEqual(2, len(profiles))
        self.assertIn(john_doe, profiles)
        self.assertIn(emma_jackson, profiles)

    def test_do_import_no_id_metadata(self):
        # given:
        row_template = MagicMock('row_template')

        # and:
        paper_metadata = MetadataEntity(content={'product_name': 'paper'},
                                        links={'delivery': ['123', '456']})
        pen_metadata = MetadataEntity(content={'product_name': 'pen'},
                                      links={'delivery': ['789']})
        row_template.do_import = MagicMock(side_effect=[paper_metadata, pen_metadata])

        # and:
        mock_template_manager = MagicMock('template_manager')
        mock_template_manager.create_row_template = MagicMock(return_value=row_template)
        mock_template_manager.get_header_row = MagicMock(return_value=['header1', 'header2'])
        mock_template_manager.get_concrete_type = MagicMock(return_value='concrete_entity')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('product')
        worksheet['A6'] = 'paper'
        worksheet['A7'] = 'pen'

        # when:
        worksheet_importer = WorksheetImporter(mock_template_manager)
        results = worksheet_importer.do_import(worksheet)

        # then:
        self.assertEqual(2, len(results))
        self.assertIn(paper_metadata, results)
        self.assertIn(pen_metadata, results)

        # and: object id should be assigned
        paper_id = paper_metadata.object_id
        self.assertIsNotNone(paper_id)
        pen_id = pen_metadata.object_id
        self.assertIsNotNone(pen_id)
        self.assertNotEqual(paper_id, pen_id)
