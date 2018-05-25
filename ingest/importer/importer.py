import re

import openpyxl

from ingest.importer.conversion import template_manager, conversion_strategy
from ingest.importer.conversion.template_manager import TemplateManager
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook
from ingest.importer.submission import IngestSubmitter, EntitiesDictionaries, EntityLinker


class IngestImporter:

    def __init__(self, ingest_api):
        self.ingest_api = ingest_api

    def import_spreadsheet(self, file_path, submission_url, dry_run=False):
        workbook = openpyxl.load_workbook(filename=file_path)
        ingest_workbook = IngestWorkbook(workbook)
        schemas = ingest_workbook.get_schemas()

        template_mgr = template_manager.build(schemas)
        workbook_importer = WorkbookImporter(template_mgr)
        spreadsheet_json = workbook_importer.do_import(ingest_workbook)

        entities_dictionaries = EntitiesDictionaries(spreadsheet_json)
        entity_linker = EntityLinker(template_mgr)
        entities_dictionaries = entity_linker.process_links(entities_dictionaries)

        submission = None
        if not dry_run:
            submitter = IngestSubmitter(self.ingest_api, template_mgr)
            submission = submitter.submit(entities_dictionaries, submission_url)
            print(f'Submission in {submission_url} is done!')

        return submission


class WorkbookImporter:

    def __init__(self, template_mgr):
        self.worksheet_importer = WorksheetImporter()
        self.template_mgr = template_mgr

    def do_import(self, workbook: IngestWorkbook):
        pre_ingest_json_map = {}

        self.import_project(pre_ingest_json_map, workbook)

        for worksheet in workbook.importable_worksheets():
            concrete_entity = self.template_mgr.get_concrete_entity_of_tab(worksheet.title)

            # TODO what if the tab is not a valid entity?
            if concrete_entity is None:
                print(f'{worksheet.title} is not a valid tab name.')
                continue

            domain_entity = self.template_mgr.get_domain_entity(concrete_entity)

            if domain_entity is None:
                continue

            entities_dict = self.worksheet_importer.do_import(worksheet, self.template_mgr)
            if pre_ingest_json_map.get(domain_entity) is None:
                pre_ingest_json_map[domain_entity] = {}

            pre_ingest_json_map[domain_entity].update(entities_dict)
        return pre_ingest_json_map

    def import_project(self, pre_ingest_json_map, workbook):
        project_worksheet = workbook.get_project_worksheet()
        project_importer = ProjectWorksheetImporter()
        project_dict = project_importer.do_import(project_worksheet, self.template_mgr)
        contact_worksheet = workbook.get_contact_worksheet()
        contact_importer = ContactWorksheetImporter()
        contacts = contact_importer.do_import(contact_worksheet, self.template_mgr)
        project_record = list(project_dict.values())[0]
        project_record['content']['contributors'] = list(map(lambda record: record['content']['contributors'][0], contacts))
        pre_ingest_json_map['project'] = project_dict


class WorksheetImporter:
    KEY_HEADER_ROW_IDX = 4
    USER_FRIENDLY_HEADER_ROW_IDX = 2
    START_ROW_IDX = 5

    UNKNOWN_ID_PREFIX = '_unknown_'

    def __init__(self):
        self.unknown_id_ctr = 0

    def do_import(self, worksheet, template: TemplateManager):
        return self._import_records(worksheet, template)

    def _import_records(self, worksheet, template: TemplateManager):
        records = {}
        row_template = template.create_row_template(worksheet)
        for row in self._get_data_rows(worksheet):
            # TODO row_template.do_import should return a structured abstraction
            json = row_template.do_import(row)

            link_map = json.get(conversion_strategy.LINKS_FIELD, {})
            new_link_map = {}

            for concrete_entity, ids in link_map.items():
                domain_entity = template.get_domain_entity(concrete_entity)

                if domain_entity is None:
                    continue

                domain_entity_ids = new_link_map.get(domain_entity, [])

                if len(domain_entity_ids) == 0:
                    new_link_map[domain_entity] = domain_entity_ids

                domain_entity_ids.extend(ids)

            concrete_entity = template.get_concrete_entity_of_tab(worksheet.title)

            json[conversion_strategy.CONTENT_FIELD]['describedBy'] = template.get_schema_url(concrete_entity)
            json[conversion_strategy.CONTENT_FIELD]['schema_type'] = template.get_domain_entity(concrete_entity)

            record_id = json.get(conversion_strategy.OBJECT_ID_FIELD, self._generate_id())

            records[record_id] = {
                'content': json[conversion_strategy.CONTENT_FIELD],
                'links_by_entity': new_link_map
            }
        return records

    def _generate_id(self):
        self.unknown_id_ctr = self.unknown_id_ctr + 1
        return f'{self.UNKNOWN_ID_PREFIX}{self.unknown_id_ctr}'

    def _get_data_rows(self, worksheet):
        return worksheet.iter_rows(row_offset=self.START_ROW_IDX,
                                   max_row=(worksheet.max_row - self.START_ROW_IDX))


class ProjectWorksheetImporter(WorksheetImporter):

    def do_import(self, worksheet, template: TemplateManager):
        records = self._import_records(worksheet, template)

        if len(records.keys()) == 0:
            raise NoProjectFound()

        if len(records.keys()) > 1:
            raise MultipleProjectsFound()

        return records


class ContactWorksheetImporter(WorksheetImporter):

    def do_import(self, worksheet, template: TemplateManager):
        records = self._import_records(worksheet, template)

        return list(records.values())

class MultipleProjectsFound(Exception):
    pass


class NoProjectFound(Exception):
    pass
